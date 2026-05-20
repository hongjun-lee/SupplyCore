#!/usr/bin/env python3
"""
add_header_comments.py — 给数据采集 xlsx 的 R1 表头单元格灌入必填字段批注（hover 提示）

背景：
    V0.2.1 ④ 总览记录"6 模板 19 sheet 累计 84 条必填字段表头批注"，但实际
    只有 07-组织架构参考表的 5 条生效；01-06 全部为 0 条。
    本脚本按 V0.2.5 用户选定的 **B 方案**（核心优先 / 02+04+06）+
    **L 方案**（每条 6 行：英文名 / 类型(长度) / ✅ 必填 / 校验 / 示例值 /
    常见错误）补齐 25 条批注。

覆盖范围（B 方案）：
    - 02 物料分类     (4 条必填)
    - 02 物料主数据   (5 条必填)
    - 02 计量单位     (3 条必填)
    - 04 仓库         (6 条必填)
    - 06 期初库存     (7 条必填)
    合计 25 条

幂等保证：
    每次运行先清空目标列已有 comment 再重写。

使用方法：
    python3 scripts/add_header_comments.py
    python3 scripts/add_header_comments.py --only 02-物资主数据模板-V0.2.xlsx
    python3 scripts/add_header_comments.py --dry-run
"""

from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List

try:
    from openpyxl import load_workbook
    from openpyxl.comments import Comment
except ImportError:
    print("错误：需要安装 openpyxl 库", file=sys.stderr)
    print("运行：pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ============================================================================
DEFAULT_DIR = "docs/上线/word/数据采集模板-xlsx"
COMMENT_AUTHOR = "SupplyCore"
COMMENT_WIDTH = 320   # 像素
COMMENT_HEIGHT = 180


# ============================================================================
# 批注内容：6 行格式
#   ① 字段英文名
#   ② 类型 (长度)
#   ③ ✅ 必填
#   ④ 校验/枚举说明
#   ⑤ 示例值
#   ⑥ 常见错误（❌ 开头）
# ============================================================================
@dataclass(frozen=True)
class CommentSpec:
    field_en: str
    type_len: str
    validation: str
    example: str
    common_error: str

    def to_comment_text(self) -> str:
        return (
            f"字段：{self.field_en}\n"
            f"类型：{self.type_len}\n"
            f"必填：✅\n"
            f"校验：{self.validation}\n"
            f"示例：{self.example}\n"
            f"❌ 常见错误：{self.common_error}"
        )


# ============================================================================
# 内置：文件 → sheet → 列中文表头 → CommentSpec
# Key 用 R1 表头实际中文（与 xlsx 一致，确保精确匹配）
# ============================================================================
COMMENTS: Dict[str, Dict[str, Dict[str, CommentSpec]]] = {
    "02-物资主数据模板-V0.2.xlsx": {
        "物料分类": {
            "分类编码": CommentSpec(
                "category_code", "string (8)",
                "V1.8 标准编码；一级如 ZH，二级如 ZH01；xlsx 已预填 110 行",
                "ZH01 / HG01 / SB10",
                "不要带横杠（写 ZH-01）；不要继续用旧 M-XX 或 L-Z-A 编码",
            ),
            "分类名称": CommentSpec(
                "category_name", "string (64)",
                "对应 category_code 的中文显示名",
                "锚杆 / 炸药 / 采煤机",
                "不要写英文；不要带具体规格",
            ),
            "分类层级": CommentSpec(
                "level", "int",
                "枚举 1=一级大类 / 2=二级分类",
                "1 或 2",
                "不要写中文「一/二」；不要写 0 或 3+",
            ),
            "是否高敏感": CommentSpec(
                "is_high_sensitive", "bool",
                "火工品(HG) / 化工材料(HX) / 燃油(YZ01) 等需 4 步走管控的分类",
                "true 或 false",
                "不要写中文「是/否」；HG/HX 大类必须 true",
            ),
        },
        "物料主数据": {
            "原系统物料编码": CommentSpec(
                "legacy_code", "string (64)",
                "迁移主键，唯一；直接取原系统物料编码原值",
                "000123 / 000456 / 000789",
                "不要填新生成的 material_code（如 ZH01000001）；不要重复",
            ),
            "原系统分类编码": CommentSpec(
                "original_category_code", "string (64)",
                "旧系统分类编码原值；系统通过 M-18 映射到 V1.8 二级分类",
                "L-Z-A-01-01-01 / OLD-HG-001",
                "不要直接填 V1.8 新编码；走映射不要手动转换",
            ),
            "物料名称": CommentSpec(
                "material_name", "string (128)",
                "物料中文名（不含规格）",
                "树脂锚杆 / 乳化炸药 / 矿用电缆",
                "不要把规格写进名称（树脂锚杆Φ20 ✗）；不要用简称",
            ),
            "规格型号": CommentSpec(
                "specification", "string (256)",
                "完整规格描述，含尺寸/型号/技术参数",
                "Φ20×2200 / 32mm×200mm / YJV22-3×95+1×50",
                "不要与物料名称重复；不要留空写「无」",
            ),
            "计量单位": CommentSpec(
                "unit", "string (16)",
                "⬇️ xlsx 已加下拉验证，从「计量单位」sheet 25 个标准单位选",
                "件 / 卷 / KG / T / M / M3",
                "不要手填非标单位（如「条」「捆」非标）；下拉外的值会被拒",
            ),
        },
        "计量单位": {
            "单位编码": CommentSpec(
                "unit_code", "string",
                "唯一；推荐使用国际单位代号或单字中文",
                "M / KG / T / 件 / 卷",
                "不要重复；不要用空格或特殊符号",
            ),
            "单位名称": CommentSpec(
                "unit_name", "string",
                "中文显示名",
                "米 / 千克 / 吨 / 件 / 卷",
                "不要填英文（如 meter）；不要与 unit_code 完全相同",
            ),
            "单位类型": CommentSpec(
                "unit_type", "enum",
                "枚举：长度 / 重量 / 体积 / 数量 / 其他",
                "长度 / 重量",
                "不要写英文（如 length）；不要自创类型",
            ),
        },
    },
    "04-仓储基础数据模板-V0.2.xlsx": {
        "仓库": {
            "仓库编码": CommentSpec(
                "warehouse_code", "string (32)",
                "唯一；按编码规则 WH-{org_code}-{seq}",
                "WH-FK002-001 / WH-FK001-001",
                "不要重复；不要用纯数字；不要遗漏前缀 WH-",
            ),
            "仓库名称": CommentSpec(
                "warehouse_name", "string (128)",
                "中文名，含矿名 + 仓库类型",
                "艾友矿主仓库 / 本部综合仓",
                "不要用简称（仅写「主仓」）；不要遗漏矿名",
            ),
            "组织编码": CommentSpec(
                "org_code", "string (32)",
                "必须在 07-组织架构参考表 / 厂矿级或部门级；可填 Catio 全编码或别名",
                "001.007.001 / 001.007.002 / FK001",
                "不要用 mock 矿名（艾友/东梁/五龙/新邱）；不要写中文矿名",
            ),
            "仓库类型": CommentSpec(
                "warehouse_type", "enum (32)",
                "枚举：主仓 / 临时仓 / 工地仓 / 库外保管 / 报废仓",
                "主仓 / 临时仓",
                "不要写英文（main 等）；火工品仓选「临时仓」+ 启用批次/有效期",
            ),
            "仓库管理员工号": CommentSpec(
                "manager_employee_no", "string (32)",
                "必须在 01 模板 User sheet + role 含「仓储」",
                "FK0010 / FK0011",
                "不要填姓名（张三 ✗）；不要遗漏（无管理员仓库会记台账）",
            ),
            "启用日期": CommentSpec(
                "active_date", "date",
                "ISO 格式 YYYY-MM-DD；仓库正式启用日期",
                "2026-05-20 / 2025-12-01",
                "不要用 2026/5/20 或 May 20；不要填未来日期",
            ),
        },
    },
    "06-期初库存模板-V0.2.xlsx": {
        "期初库存": {
            "物料编码": CommentSpec(
                "material_code", "string (64)",
                "可填新 material_code（如 ZH01000001）或旧 legacy_code（如 000123），系统两路反查",
                "ZH01000001 / 000123",
                "不要带横杠（写 ZH-01-000001 ✗）；不要混填两种格式同一行",
            ),
            "仓库编码": CommentSpec(
                "warehouse_code", "string (32)",
                "必须在 04 模板 Warehouse sheet（先填 04 再填 06）",
                "WH-FK002-001",
                "不要新建未注册仓库；不要拼写错误",
            ),
            "数量": CommentSpec(
                "quantity", "decimal",
                "盘点实际数量；必须 > 0",
                "500 / 17750.5",
                "不要填 0 或负数；不要带单位（写 500 件 ✗）",
            ),
            "计量单位": CommentSpec(
                "unit", "string (16)",
                "必须在 02 计量单位 sheet + 与 Material.unit 一致",
                "件 / 卷 / KG / M",
                "不要与物料定义的 unit 不一致（会触发 stock_unit_mismatch 台账）",
            ),
            "入库日期": CommentSpec(
                "inbound_date", "date",
                "期初库存对应的实际入库日期；YYYY-MM-DD",
                "2025-12-01 / 2026-04-20",
                "不要填未来日期；不要全部填一天（不真实）",
            ),
            "填报日期": CommentSpec(
                "report_date", "date",
                "本次盘点填报当日；YYYY-MM-DD",
                "2026-05-20",
                "不要填一年前；与 inbound_date 区分（inbound 是入库当日）",
            ),
            "数据来源": CommentSpec(
                "source", "enum (32)",
                "枚举：线下台账 / 上一系统迁移 / 物理盘点",
                "物理盘点 / 线下台账",
                "不要写英文；不要自创来源（如「估算」）",
            ),
        },
    },
}


# ============================================================================
# 业务函数
# ============================================================================
def process_sheet(ws, sheet_specs: Dict[str, CommentSpec]) -> tuple[int, List[str]]:
    """给 sheet 的 R1 表头按 sheet_specs 加批注。返回 (写入数, 未匹配列名列表)。"""
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if isinstance(v, str):
            headers[v] = c

    written = 0
    not_found = []
    for header_zh, spec in sheet_specs.items():
        if header_zh not in headers:
            not_found.append(header_zh)
            continue
        col = headers[header_zh]
        cell = ws.cell(row=1, column=col)
        comment = Comment(spec.to_comment_text(), COMMENT_AUTHOR)
        comment.width = COMMENT_WIDTH
        comment.height = COMMENT_HEIGHT
        cell.comment = comment  # 直接覆盖既有 comment 实现幂等
        written += 1

    return written, not_found


@dataclass
class FileResult:
    file: str
    status: str  # 'added' / 'skipped' / 'error'
    detail: str


def process_file(path: Path, dry_run: bool, only_sheet: str | None = None) -> FileResult:
    fname = path.name
    if fname not in COMMENTS:
        return FileResult(fname, "skipped", "不在内置批注映射列表")

    lock = path.parent / f".~{fname}"
    if lock.exists() and not dry_run:
        return FileResult(fname, "error", f"检测到锁文件 {lock.name}，请关闭 Excel/WPS")

    try:
        wb = load_workbook(path)
    except Exception as e:
        if dry_run:
            return FileResult(fname, "added", f"(dry-run) 计划写入 {sum(len(s) for s in COMMENTS[fname].values())} 条批注")
        return FileResult(fname, "error", f"无法打开 xlsx：{e}")

    total_written = 0
    sheet_reports = []
    for sn, specs in COMMENTS[fname].items():
        if only_sheet and sn != only_sheet:
            continue
        if sn not in wb.sheetnames:
            sheet_reports.append(f"{sn}=❌缺失")
            continue
        if dry_run:
            sheet_reports.append(f"{sn}={len(specs)} 条")
            total_written += len(specs)
            continue
        ws = wb[sn]
        n, nf = process_sheet(ws, specs)
        total_written += n
        sheet_reports.append(f"{sn}={n} 条" + (f"(未匹配:{nf})" if nf else ""))

    if not dry_run:
        wb.save(path)
    wb.close()

    return FileResult(fname, "added", f"{total_written} 条 — " + " / ".join(sheet_reports))


# ============================================================================
def find_repo_root(start: Path) -> Path:
    p = start.resolve()
    for parent in [p, *p.parents]:
        if (parent / ".git").exists():
            return parent
    return start.parent


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="给 02/04/06 数据采集 xlsx 的 R1 表头加 ✅ 必填字段批注（B+L 方案 / 25 条）",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--dir", type=str, default=None,
                        help=f"模板目录（默认仓库内 {DEFAULT_DIR}）")
    parser.add_argument("--only", type=str, default=None,
                        help="只处理指定文件名")
    parser.add_argument("--sheet", type=str, default=None,
                        help="只处理指定 sheet 名（需与 --only 配合）")
    parser.add_argument("--dry-run", action="store_true",
                        help="不写盘，只打印")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    repo_root = find_repo_root(Path(__file__))
    target_dir = Path(args.dir).expanduser().resolve() if args.dir else repo_root / DEFAULT_DIR

    if not target_dir.exists():
        print(f"错误：目录不存在：{target_dir}", file=sys.stderr)
        return 2

    files = sorted(COMMENTS.keys())
    if args.only:
        if args.only not in COMMENTS:
            print(f"错误：--only {args.only} 不在内置批注列表；可选：", file=sys.stderr)
            for f in files:
                print(f"  - {f}", file=sys.stderr)
            return 2
        files = [args.only]

    print(f"[add_header_comments] 目录：{target_dir}")
    print(f"[add_header_comments] 待处理：{len(files)} 个文件  dry_run={args.dry_run}")
    print()

    results: List[FileResult] = []
    for fname in files:
        path = target_dir / fname
        if not path.exists():
            results.append(FileResult(fname, "error", "文件不存在"))
            continue
        results.append(process_file(path, dry_run=args.dry_run, only_sheet=args.sheet))

    counts = {"added": 0, "skipped": 0, "error": 0}
    for r in results:
        counts[r.status] += 1
        icon = {"added": "+", "skipped": "·", "error": "x"}[r.status]
        print(f"  [{icon}] {r.file}: {r.detail}")

    print()
    print(f"汇总：added={counts['added']} / skipped={counts['skipped']} / error={counts['error']}")
    return 0 if counts["error"] == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
