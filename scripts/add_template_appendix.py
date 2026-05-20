#!/usr/bin/env python3
"""
add_template_appendix.py — 给数据采集 xlsx 的「说明」sheet 末尾追加
                          "附录：Sheet ↔ 数据表对照" 3 列表格

背景：
    V0.2.1 总览记录说"6 份 xlsx 都追加了附录"，但实际只有 07-组织架构
    参考表加成功了；01-06 六份并未生效。本脚本一次性补齐。

格式（与 07 xlsx R43:R46 完全一致，便于风格统一）：
    R(N):       大标题   "附录：Sheet ↔ 数据表对照"（A 列，加粗）
    R(N+1):     表头 3 列 [Sheet 中文名 | 数据表英文名 | 用途说明]
                          蓝底 FF4472C4 / 白字 / 微软雅黑 11pt bold
    R(N+2+):    数据行 3 列，细边框 + 微软雅黑 10pt

幂等保证：
    每次运行先检测「说明」sheet 是否已含"附录：Sheet ↔ 数据表对照"标记行，
    如有则截断至标记行前重写，保证多次运行结果一致。

使用方法：
    # 默认处理 01-06 六份 xlsx
    python3 scripts/add_template_appendix.py

    # 处理单个文件
    python3 scripts/add_template_appendix.py --only 02-物资主数据模板-V0.2.xlsx

    # dry-run 只打印计划
    python3 scripts/add_template_appendix.py --dry-run
"""

from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError:
    print("错误：需要安装 openpyxl 库", file=sys.stderr)
    print("运行：pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ============================================================================
# 默认目录
# ============================================================================
DEFAULT_DIR = "docs/上线/word/数据采集模板-xlsx"

# ============================================================================
# 内置映射表 — 每个 xlsx 的 sheet → 数据表 → 用途
# ----------------------------------------------------------------------------
# Key   = xlsx 文件名（不含目录）
# Value = [(sheet 中文名, 数据表英文名, 用途说明), ...]
#
# 表名对齐口径：
#   - 02/03/04/05/06：直接采自各模板 md 章节标题 `(Sheet N: EnglishName)`
#   - 01：sheet 名（A/B/C）已与 md 章节脱钩（V0.2.1 §⑤），按业务实际用途填
#   - 07：本脚本不动（07 已在创建时同步落地附录）
# ============================================================================
APPENDIX_TITLE = "附录：Sheet ↔ 数据表对照"
APPENDIX_HEADERS = ("Sheet 中文名", "数据表英文名", "用途说明")

MAPPINGS: Dict[str, List[Tuple[str, str, str]]] = {
    "01-组织与人员模板-V0.2.xlsx": [
        ("说明",          "—（无对应 DB 表）",          "本模板使用说明 + 附录对照"),
        ("A_业务联系人",  "OrganizationContact (业务台账)", "A 场景：厂矿/部门主要业务联系人（Nova 缺失时手填）"),
        ("B_数据责任人",  "DataOwner (业务台账)",        "B 场景：各类基础数据采集的责任人台账"),
        ("C_Nova异常申报", "NovaIssueReport (提报票据)", "C 场景：Nova 主数据缺失/错误申报"),
    ],
    "02-物资主数据模板-V0.2.xlsx": [
        ("说明",            "—（无对应 DB 表）",            "本模板使用说明 + 附录对照"),
        ("物料分类",        "MaterialCategory",             "V1.8 14 启用 + 1 保留大类、95 个二级分类基线（已预填）"),
        ("原分类映射(M-18)", "MaterialCategoryLegacyMapping", "原系统旧分类 ↔ V1.8 新二级分类映射（M-18 关键台账）"),
        ("物料主数据",      "Material",                     "物料主数据；填 legacy_code + original_category_code，新 material_code 自动生成"),
        ("计量单位",        "Unit",                         "25 个常用标准单位字典（已预填）"),
    ],
    "03-供应商档案模板-V0.2.xlsx": [
        ("说明",          "—（无对应 DB 表）",      "本模板使用说明 + 附录对照"),
        ("供应商",        "Supplier",               "供应商基础档案"),
        ("供应商联系人",  "SupplierContact",        "供应商联系人"),
        ("供应商银行账户", "SupplierBank",          "供应商开户/银行账户"),
        ("供应商资质",    "SupplierQualification",  "供应商资质证照（营业执照、安许等）"),
        ("NC供应商映射",  "NcSupplierMapping",      "SupplyCore ↔ NC 供应商编码映射"),
    ],
    "04-仓储基础数据模板-V0.2.xlsx": [
        ("说明", "—（无对应 DB 表）",  "本模板使用说明 + 附录对照"),
        ("仓库", "Warehouse",          "仓库基础档案"),
        ("库区", "WarehouseZone",      "仓库内库区划分"),
        ("货位", "StorageLocation",    "库区内货位/库位"),
    ],
    "05-财务与NC映射模板-V0.2.xlsx": [
        ("说明",             "—（无对应 DB 表）",      "本模板使用说明 + 附录对照"),
        ("部门映射",         "DepartmentMapping",      "部门编码映射（SupplyCore ↔ NC）"),
        ("供应商映射",       "SupplierMapping",        "供应商编码映射（财务侧视角；与 03 NC 供应商映射协同）"),
        ("科目映射",         "AccountMapping",         "会计科目映射（SupplyCore 业务事件 → NC 科目）"),
        ("接口编码映射",     "InterfaceCodeMapping",   "业务单据类型与 NC InterfaceCode 映射"),
        ("凭证号规则(参考)", "VoucherNumberRule",      "SC/NC 双号制规则（仅参考，不需填报）"),
    ],
    "06-期初库存模板-V0.2.xlsx": [
        ("说明",     "—（无对应 DB 表）", "本模板使用说明 + 附录对照"),
        ("期初库存", "InitialStock",       "期初库存盘点数据（上线切换基础）"),
    ],
}

# ============================================================================
# 样式（对齐 07 xlsx）
# ============================================================================
THIN = Side(border_style="thin", color="B4B4B4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FONT_TITLE = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
FONT_HEADER = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="微软雅黑", size=10)

FILL_TITLE = PatternFill("solid", fgColor="305496")   # 深蓝（标题）
FILL_HEADER = PatternFill("solid", fgColor="4472C4")  # 蓝（表头，对齐 07）

ALIGN_LEFT = Alignment(vertical="center", horizontal="left", wrap_text=True)
ALIGN_CENTER = Alignment(vertical="center", horizontal="center", wrap_text=True)

COL_WIDTHS = (20, 28, 68)  # A / B / C


# ============================================================================
# 业务函数
# ============================================================================
@dataclass
class FileResult:
    file: str
    status: str   # 'added' / 'updated' / 'skipped' / 'error'
    detail: str


def find_appendix_start_row(ws) -> int | None:
    """检测「说明」sheet 中已有附录标题行的位置；找不到返回 None。"""
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and APPENDIX_TITLE in v:
            return r
    return None


def truncate_appendix(ws, start_row: int) -> None:
    """删除附录起始行及之后所有行（含可能的空行）。"""
    if start_row <= ws.max_row:
        # 多删一些以清掉附录前可能留的空行（safe upper bound）
        ws.delete_rows(start_row, ws.max_row - start_row + 1)


def append_appendix(ws, rows: List[Tuple[str, str, str]]) -> Tuple[int, int]:
    """
    在「说明」sheet 末尾追加附录。返回 (appendix_title_row, last_data_row)。

    布局：
        [last_existing_row]     ← 现有内容末行
        +1 空行
        +1 标题（深蓝底白字 / 12pt bold，3 列合并）
        +1 表头（蓝底白字 / 11pt bold）
        +N 数据行（边框 + 10pt）
    """
    # 现有非空末行
    last_existing = ws.max_row
    while last_existing >= 1:
        if any(
            ws.cell(row=last_existing, column=c).value not in (None, "")
            for c in range(1, max(ws.max_column, 1) + 1)
        ):
            break
        last_existing -= 1

    title_row = last_existing + 2

    # 标题
    ws.cell(row=title_row, column=1, value=APPENDIX_TITLE)
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=3)
    title_cell = ws.cell(row=title_row, column=1)
    title_cell.font = FONT_TITLE
    title_cell.fill = FILL_TITLE
    title_cell.alignment = ALIGN_LEFT
    title_cell.border = BORDER
    # 给合并区其他单元格也填上边框/背景，避免视觉断裂
    for c in (2, 3):
        cell = ws.cell(row=title_row, column=c)
        cell.fill = FILL_TITLE
        cell.border = BORDER

    # 表头
    header_row = title_row + 1
    for c, header in enumerate(APPENDIX_HEADERS, start=1):
        cell = ws.cell(row=header_row, column=c, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER

    # 数据
    last_data_row = header_row
    for i, (sn, tn, purpose) in enumerate(rows, start=1):
        r = header_row + i
        ws.cell(row=r, column=1, value=sn).font = FONT_BODY
        ws.cell(row=r, column=2, value=tn).font = FONT_BODY
        ws.cell(row=r, column=3, value=purpose).font = FONT_BODY
        for c in range(1, 4):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = ALIGN_CENTER if c in (1, 2) else ALIGN_LEFT
        last_data_row = r

    # 列宽
    for col_letter, width in zip("ABC", COL_WIDTHS):
        if ws.column_dimensions[col_letter].width is None or ws.column_dimensions[col_letter].width < width:
            ws.column_dimensions[col_letter].width = width

    return title_row, last_data_row


def process_file(path: Path, dry_run: bool) -> FileResult:
    fname = path.name
    if fname not in MAPPINGS:
        return FileResult(fname, "skipped", "不在内置映射列表（07/其他文件本脚本不处理）")

    lock = path.parent / f".~{fname}"
    if lock.exists() and not dry_run:
        return FileResult(fname, "error", f"检测到锁文件 {lock.name}，请关闭 Excel/WPS")

    try:
        wb = load_workbook(path)
    except Exception as e:
        if dry_run:
            # dry-run 容错：文件被锁也能给出计划
            existing_start_hint = "(无法读取文件，假定未存在附录)"
            return FileResult(
                fname,
                "added",
                f"将在「说明」sheet 末尾写入 {len(MAPPINGS[fname])} 行数据 {existing_start_hint}",
            )
        return FileResult(fname, "error", f"无法打开 xlsx：{e}")

    if "说明" not in wb.sheetnames:
        return FileResult(fname, "error", "xlsx 缺少「说明」sheet")

    ws = wb["说明"]
    existing_start = find_appendix_start_row(ws)
    will_be = "updated" if existing_start else "added"

    rows = MAPPINGS[fname]

    if dry_run:
        return FileResult(
            fname,
            will_be,
            f"将在「说明」sheet 末尾写入 {len(rows)} 行数据"
            + (f"（先截断已有附录 @ R{existing_start}）" if existing_start else ""),
        )

    if existing_start:
        truncate_appendix(ws, existing_start)

    title_row, last_data_row = append_appendix(ws, rows)
    wb.save(path)
    wb.close()

    return FileResult(
        fname,
        will_be,
        f"附录 @ R{title_row} 标题 / R{title_row + 1} 表头 / R{title_row + 2}-R{last_data_row} 数据 ({len(rows)} 行)",
    )


# ============================================================================
# CLI
# ============================================================================
def find_repo_root(start: Path) -> Path:
    p = start.resolve()
    for parent in [p, *p.parents]:
        if (parent / ".git").exists():
            return parent
    return start.parent


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="给 01-06 数据采集 xlsx 的「说明」sheet 追加 'Sheet ↔ 数据表对照' 附录",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "--dir",
        type=str,
        default=None,
        help=f"模板目录（默认仓库内 {DEFAULT_DIR}）",
    )
    parser.add_argument(
        "--only",
        type=str,
        default=None,
        help="只处理指定文件名（不含目录），如 02-物资主数据模板-V0.2.xlsx",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="不写盘，只打印将要做的事",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    repo_root = find_repo_root(Path(__file__))
    target_dir = Path(args.dir).expanduser().resolve() if args.dir else repo_root / DEFAULT_DIR

    if not target_dir.exists():
        print(f"错误：目录不存在：{target_dir}", file=sys.stderr)
        return 2

    files = sorted(MAPPINGS.keys())
    if args.only:
        if args.only not in MAPPINGS:
            print(f"错误：--only {args.only} 不在内置映射列表；可选：", file=sys.stderr)
            for f in files:
                print(f"  - {f}", file=sys.stderr)
            return 2
        files = [args.only]

    print(f"[add_template_appendix] 目录：{target_dir}")
    print(f"[add_template_appendix] 待处理：{len(files)} 个文件  dry_run={args.dry_run}")
    print()

    results: List[FileResult] = []
    for fname in files:
        path = target_dir / fname
        if not path.exists():
            results.append(FileResult(fname, "error", "文件不存在"))
            continue
        results.append(process_file(path, dry_run=args.dry_run))

    # 输出汇总
    counts = {"added": 0, "updated": 0, "skipped": 0, "error": 0}
    for r in results:
        counts[r.status] += 1
        icon = {"added": "+", "updated": "~", "skipped": "·", "error": "x"}[r.status]
        print(f"  [{icon}] {r.file}: {r.detail}")

    print()
    print(f"汇总：added={counts['added']} / updated={counts['updated']} / skipped={counts['skipped']} / error={counts['error']}")
    return 0 if counts["error"] == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
