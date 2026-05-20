#!/usr/bin/env python3
"""
add_validations.py — 给 04/06 数据采集 xlsx 加同 sheet 数据验证（S1 范围 / 9 处）

背景：
    继 V0.2.3 (02 跨 sheet DV) + V0.2.5 (formula1 不带 = 修复) 之后，
    用户选定 S1 范围：04+06 共 9 处同 sheet 数据验证（0 跨表风险）。

覆盖（9 处 = 5 A 类枚举 + 4 C 类数字/日期）：
    04/仓库/D    enum  主仓/临时仓/工地仓/库外保管/报废仓
    04/仓库/H    date  active_date ≤ TODAY()
    04/库区/D    enum  普通库区/危险品库区/低温库区/隔离区/待检区
    04/货位/H    enum  可用/冻结/维护/报废
    06/期初库存/H  decimal  quantity > 0
    06/期初库存/L  date     inbound_date ≤ TODAY()
    06/期初库存/Q  enum     正常/待检/隔离/报废
    06/期初库存/S  date     report_date ∈ [TODAY()-7, TODAY()+7]
    06/期初库存/T  enum     线下台账/上一系统迁移/物理盘点

关键经验（从 V0.2.5 痛中学到）：
    - openpyxl DataValidation.formula1 **不带前导 `=`**（带了 Excel/WPS 会
      静默丢弃 DV 甚至自动修复触发数据回退）。
    - 同 sheet enum list 的 formula1 格式是 `"v1,v2,v3"`（**带双引号包裹**）。
    - date/decimal 范围 DV 用 `formula1` + 可选 `formula2`，operator 取
      lessThanOrEqual / greaterThan / between / 等。
    - 写入后必须 unzip 看 xml 才算真正验证（openpyxl 自身能读非法 xml）。

幂等：
    每次运行先清掉目标列已有 list/date/decimal DV，再重写。

用法：
    python3 scripts/add_validations.py
    python3 scripts/add_validations.py --only 04-仓储基础数据模板-V0.2.xlsx
    python3 scripts/add_validations.py --dry-run
"""

from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Literal

try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("错误：需要安装 openpyxl 库", file=sys.stderr)
    print("运行：pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ============================================================================
DEFAULT_DIR = "docs/上线/word/数据采集模板-xlsx"
DV_RANGE_END = 2000  # 业务方可填的最大行数


@dataclass
class Rule:
    """单条数据验证规则。"""
    col_letter: str          # 'D' / 'H' / 'L' / ...
    label: str               # 描述用，便于日志
    dv_type: Literal["list", "decimal", "date", "whole"]
    formula1: str            # ⚠️ 不带前导 `=`
    formula2: str | None = None
    operator: str | None = None  # equal / notEqual / between / greaterThan / lessThan / greaterThanOrEqual / lessThanOrEqual
    prompt: str = ""
    error: str = ""
    error_title: str = "数据不合规范"
    prompt_title: str = ""

    def build(self) -> DataValidation:
        kwargs = dict(
            type=self.dv_type,
            formula1=self.formula1,
            allow_blank=True,
            showErrorMessage=True,
            errorTitle=self.error_title,
            error=self.error,
            showInputMessage=bool(self.prompt),
            promptTitle=self.prompt_title or self.label,
            prompt=self.prompt,
        )
        if self.formula2 is not None:
            kwargs["formula2"] = self.formula2
        if self.operator is not None:
            kwargs["operator"] = self.operator
        dv = DataValidation(**kwargs)
        dv.add(f"{self.col_letter}2:{self.col_letter}{DV_RANGE_END}")
        return dv


# ============================================================================
# S1 严格范围：04+06 共 9 处
# ============================================================================
VALIDATIONS: Dict[str, Dict[str, List[Rule]]] = {
    "04-仓储基础数据模板-V0.2.xlsx": {
        "仓库": [
            Rule(
                col_letter="D", label="仓库类型",
                dv_type="list",
                formula1='"主仓,临时仓,工地仓,库外保管,报废仓"',
                error="必须从下拉中选择：主仓 / 临时仓 / 工地仓 / 库外保管 / 报废仓",
                prompt="从下拉选择仓库类型。火工品仓建议「临时仓」+ 启用批次/有效期",
            ),
            Rule(
                col_letter="H", label="启用日期",
                dv_type="date", operator="lessThanOrEqual",
                formula1="TODAY()",
                error="启用日期不能晚于今日",
                prompt="ISO 格式 YYYY-MM-DD；不晚于今日",
            ),
        ],
        "库区": [
            Rule(
                col_letter="D", label="zone_type",
                dv_type="list",
                formula1='"普通库区,危险品库区,低温库区,隔离区,待检区"',
                error="必须从下拉中选择：普通库区 / 危险品库区 / 低温库区 / 隔离区 / 待检区",
                prompt="火工品 / 易燃易爆库区选「危险品库区」+ 仓库须 enable_batch / enable_expiry",
            ),
        ],
        "货位": [
            Rule(
                col_letter="H", label="location_status",
                dv_type="list",
                formula1='"可用,冻结,维护,报废"',
                error="必须从下拉中选择：可用 / 冻结 / 维护 / 报废",
                prompt="选「冻结」时下方 freeze_reason 列必填",
            ),
        ],
    },
    "06-期初库存模板-V0.2.xlsx": {
        "期初库存": [
            Rule(
                col_letter="H", label="数量",
                dv_type="decimal", operator="greaterThan",
                formula1="0",
                error="数量必须大于 0；零库存请不要导入此行",
                prompt="盘点实际数量；> 0；不要带单位（写 500 件 ❌）",
            ),
            Rule(
                col_letter="L", label="入库日期",
                dv_type="date", operator="lessThanOrEqual",
                formula1="TODAY()",
                error="入库日期不能晚于今日；不要填未来日期",
                prompt="期初对应实际入库日期，YYYY-MM-DD；不晚于今日",
            ),
            Rule(
                col_letter="Q", label="库存状态",
                dv_type="list",
                formula1='"正常,待检,隔离,报废"',
                error="必须从下拉中选择：正常 / 待检 / 隔离 / 报废",
                prompt="非「正常」时需在调整说明列写明原因；触发 stock_state_abnormal 台账",
            ),
            Rule(
                col_letter="S", label="填报日期",
                dv_type="date", operator="between",
                formula1="TODAY()-7", formula2="TODAY()+7",
                error="填报日期应在今日 ±7 天范围内（防止 1 年前的数据误填）",
                prompt="本次盘点填报当日，YYYY-MM-DD；与入库日期区分",
            ),
            Rule(
                col_letter="T", label="数据来源",
                dv_type="list",
                formula1='"线下台账,上一系统迁移,物理盘点"',
                error="必须从下拉中选择：线下台账 / 上一系统迁移 / 物理盘点",
                prompt="标识本行库存数据的来源；财务对账时区分口径",
            ),
        ],
    },
}


# ============================================================================
def _rule_targets_column(dv: DataValidation, col_letter: str) -> bool:
    """判断 DV 是否覆盖了指定列（用于幂等清理）。"""
    try:
        return f"{col_letter}2:" in str(dv.sqref)
    except Exception:
        return False


def process_sheet(ws, rules: List[Rule]) -> int:
    """给 sheet 加 DV；先清除我们管的列上已有 DV，再重写。返回写入数。"""
    cleanup_cols = {r.col_letter for r in rules}
    existing = list(ws.data_validations.dataValidation)
    ws.data_validations.dataValidation = [
        dv for dv in existing
        if not any(_rule_targets_column(dv, col) for col in cleanup_cols)
    ]

    for r in rules:
        ws.add_data_validation(r.build())
    return len(rules)


@dataclass
class FileResult:
    file: str
    status: str  # 'added' / 'skipped' / 'error'
    detail: str


def process_file(path: Path, dry_run: bool) -> FileResult:
    fname = path.name
    if fname not in VALIDATIONS:
        return FileResult(fname, "skipped", "不在 S1 范围")

    lock = path.parent / f".~{fname}"
    if lock.exists() and not dry_run:
        return FileResult(fname, "error", f"检测到锁文件 {lock.name}，请关闭 Excel/WPS")

    try:
        wb = load_workbook(path)
    except Exception as e:
        if dry_run:
            return FileResult(fname, "added", f"(dry-run) 计划写入 {sum(len(rs) for rs in VALIDATIONS[fname].values())} 条 DV")
        return FileResult(fname, "error", f"无法打开 xlsx：{e}")

    total = 0
    sheet_reports = []
    for sn, rules in VALIDATIONS[fname].items():
        if sn not in wb.sheetnames:
            sheet_reports.append(f"{sn}=❌缺失")
            continue
        if dry_run:
            sheet_reports.append(f"{sn}={len(rules)} 条")
            total += len(rules)
            continue
        n = process_sheet(wb[sn], rules)
        total += n
        sheet_reports.append(f"{sn}={n} 条")

    if not dry_run:
        wb.save(path)
    wb.close()
    return FileResult(fname, "added", f"{total} 条 — " + " / ".join(sheet_reports))


# ============================================================================
def find_repo_root(start: Path) -> Path:
    p = start.resolve()
    for parent in [p, *p.parents]:
        if (parent / ".git").exists():
            return parent
    return start.parent


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="给 04/06 数据采集 xlsx 加同 sheet 数据验证（S1 / 9 条）",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--dir", default=None,
                        help=f"模板目录（默认仓库内 {DEFAULT_DIR}）")
    parser.add_argument("--only", default=None,
                        help="只处理指定文件名")
    parser.add_argument("--dry-run", action="store_true",
                        help="不写盘，只打印")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    repo_root = find_repo_root(Path(__file__))
    target_dir = Path(args.dir).expanduser().resolve() if args.dir else repo_root / DEFAULT_DIR
    if not target_dir.exists():
        print(f"错误：目录不存在：{target_dir}", file=sys.stderr)
        return 2

    files = sorted(VALIDATIONS.keys())
    if args.only:
        if args.only not in VALIDATIONS:
            print(f"错误：--only {args.only} 不在 S1 范围。可选：", file=sys.stderr)
            for f in files:
                print(f"  - {f}", file=sys.stderr)
            return 2
        files = [args.only]

    print(f"[add_validations] 目录：{target_dir}")
    print(f"[add_validations] 待处理：{len(files)} 个文件  dry_run={args.dry_run}")
    print()

    results = []
    for fname in files:
        path = target_dir / fname
        if not path.exists():
            results.append(FileResult(fname, "error", "文件不存在"))
            continue
        results.append(process_file(path, dry_run=args.dry_run))

    counts = {"added": 0, "skipped": 0, "error": 0}
    for r in results:
        counts[r.status] += 1
        icon = {"added": "+", "skipped": "·", "error": "x"}[r.status]
        print(f"  [{icon}] {r.file}: {r.detail}")

    print()
    print(f"汇总：added={counts['added']} / skipped={counts['skipped']} / error={counts['error']}")
    return 0 if counts["error"] == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
