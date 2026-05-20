#!/usr/bin/env python3
"""
fill_material_template.py — 02 物资主数据 xlsx 灌库脚本（V1.8 权威基线一键重生成）

用途：
    把 V1.8《物资编码规范》15 大类 + 95 二级分类、25 个标准计量单位、
    以及 Excel 下拉数据验证一次性写入 02-物资主数据模板 xlsx，业务方填
    `unit` / `category_code` 列时只能从权威列表中选择，避免乱填。

数据源：
    - 物料分类：docs/上线/物料分类映射指南-V0.1.md §二
      （与《物资编码规范》V1.8 §四 100% 一致）
    - 计量单位：内置 25 个常用单位（长度 4 / 重量 3 / 体积 3 / 数量 15）

写入目标 sheet：
    - 「物料分类」：清空 R2+ 后重新灌入 110 行（15 大类 + 95 二级）
    - 「计量单位」：清空 R2+ 后重新灌入 25 行
    - 「物料主数据」：给 F 列 unit / G 列 category_code 加 Excel 下拉验证
      （覆盖 R2:R2000）

使用方法：
    # 默认目标（02-物资主数据模板-V0.2.xlsx），含 PS 保留大类
    python3 scripts/fill_material_template.py

    # 指定其他 xlsx 路径
    python3 scripts/fill_material_template.py --file path/to/template.xlsx

    # PS 保留大类完全不进下拉（PS 区移到 sheet 末尾，下拉范围截断）
    python3 scripts/fill_material_template.py --exclude-ps

    # 不写盘，只打印将要做的事
    python3 scripts/fill_material_template.py --dry-run

注意：
    - 运行前请关闭 Excel / WPS 等可能锁定该 xlsx 的客户端
    - 脚本是幂等的：重复执行结果一致；变更前请确认 git 状态以便回滚
"""

from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import List

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("错误：需要安装 openpyxl 库", file=sys.stderr)
    print("运行：pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ============================================================================
# 默认路径（相对仓库根；脚本会自动定位仓库根）
# ============================================================================
DEFAULT_REL_PATH = "docs/上线/word/数据采集模板-xlsx/02-物资主数据模板-V0.2.xlsx"


# ============================================================================
# 数据：14 启用大类 + 1 保留大类（PS 排水材料 暂停用）
# 行顺序按当前 docs/上线/物料分类映射指南-V0.1.md §二 排列
# ============================================================================
@dataclass(frozen=True)
class Cat1:
    code: str
    name: str
    high_sensitive: bool
    is_active: bool


CATEGORIES_L1: List[Cat1] = [
    Cat1("HG", "火工品",    True,  True),
    Cat1("ZH", "支护材料",  False, True),
    Cat1("SB", "设备整机",  False, True),
    Cat1("BP", "备品备件",  False, True),
    Cat1("JD", "机电材料",  False, True),
    Cat1("YZ", "油脂燃料",  False, True),
    Cat1("GC", "钢材",      False, True),
    Cat1("JZ", "建筑材料",  False, True),
    Cat1("TF", "通防材料",  False, True),
    Cat1("HX", "化工材料",  True,  True),
    Cat1("GJ", "工器具",    False, True),
    Cat1("LB", "劳保用品",  False, True),
    Cat1("BZ", "包装材料",  False, True),
    Cat1("BG", "办公用品",  False, True),
    Cat1("PS", "排水材料(保留 / 暂停用)", False, False),
]

# ============================================================================
# 数据：95 个二级分类
# ============================================================================
@dataclass(frozen=True)
class Cat2:
    code: str
    name: str
    parent: str
    high_sensitive: bool


CATEGORIES_L2: List[Cat2] = [
    # HG 火工品 4 个 / 全部高敏感
    Cat2("HG01", "炸药",        "HG", True),
    Cat2("HG02", "雷管",        "HG", True),
    Cat2("HG03", "起爆器材",    "HG", True),
    Cat2("HG99", "其他火工品",  "HG", True),
    # ZH 支护材料 9 个
    Cat2("ZH01", "锚杆",         "ZH", False),
    Cat2("ZH02", "锚索",         "ZH", False),
    Cat2("ZH03", "钢带/托梁",    "ZH", False),
    Cat2("ZH04", "网片",         "ZH", False),
    Cat2("ZH05", "型钢支架",     "ZH", False),
    Cat2("ZH06", "坑木/方木",    "ZH", False),
    Cat2("ZH07", "托盘/锁具",    "ZH", False),
    Cat2("ZH08", "锚固材料",     "ZH", False),
    Cat2("ZH99", "其他支护材料", "ZH", False),
    # SB 设备整机 11 个
    Cat2("SB01", "采煤机",                       "SB", False),
    Cat2("SB02", "掘进机",                       "SB", False),
    Cat2("SB03", "液压支架",                     "SB", False),
    Cat2("SB04", "输送设备(皮带机、刮板机)",     "SB", False),
    Cat2("SB05", "提升设备(绞车、提升机)",       "SB", False),
    Cat2("SB06", "泵类设备(水泵、乳化液泵)",     "SB", False),
    Cat2("SB07", "压风设备",                     "SB", False),
    Cat2("SB08", "通风设备",                     "SB", False),
    Cat2("SB09", "洗选设备",                     "SB", False),
    Cat2("SB10", "运输设备(井下车辆、单轨吊等)", "SB", False),
    Cat2("SB99", "其他设备",                     "SB", False),
    # BP 备品备件 10 个
    Cat2("BP01", "采煤机配件",   "BP", False),
    Cat2("BP02", "掘进机配件",   "BP", False),
    Cat2("BP03", "液压支架配件", "BP", False),
    Cat2("BP04", "输送设备配件", "BP", False),
    Cat2("BP05", "提升设备配件", "BP", False),
    Cat2("BP06", "泵类配件",     "BP", False),
    Cat2("BP07", "压风机配件",   "BP", False),
    Cat2("BP08", "通风机配件",   "BP", False),
    Cat2("BP09", "洗选设备配件", "BP", False),
    Cat2("BP99", "其他设备配件", "BP", False),
    # JD 机电材料 6 个
    Cat2("JD01", "电缆",         "JD", False),
    Cat2("JD02", "防爆电器",     "JD", False),
    Cat2("JD03", "传感器/仪表",  "JD", False),
    Cat2("JD04", "控制设备",     "JD", False),
    Cat2("JD05", "照明器材",     "JD", False),
    Cat2("JD99", "其他机电材料", "JD", False),
    # YZ 油脂燃料 6 个（YZ01 燃油部分高敏感）
    Cat2("YZ01", "燃油",     "YZ", True),
    Cat2("YZ02", "液压油",   "YZ", False),
    Cat2("YZ03", "乳化液",   "YZ", False),
    Cat2("YZ04", "润滑油",   "YZ", False),
    Cat2("YZ05", "润滑脂",   "YZ", False),
    Cat2("YZ99", "其他油料", "YZ", False),
    # GC 钢材 8 个
    Cat2("GC01", "钢轨",      "GC", False),
    Cat2("GC02", "型钢",      "GC", False),
    Cat2("GC03", "钢板",      "GC", False),
    Cat2("GC04", "钢管",      "GC", False),
    Cat2("GC05", "钢丝绳",    "GC", False),
    Cat2("GC06", "线材/焊材", "GC", False),
    Cat2("GC07", "铸锻件",    "GC", False),
    Cat2("GC99", "其他钢材",  "GC", False),
    # JZ 建筑材料 5 个
    Cat2("JZ01", "水泥",     "JZ", False),
    Cat2("JZ02", "砂石料",   "JZ", False),
    Cat2("JZ03", "砖/砌块",  "JZ", False),
    Cat2("JZ04", "防水材料", "JZ", False),
    Cat2("JZ99", "其他建材", "JZ", False),
    # TF 通防材料 5 个
    Cat2("TF01", "风筒",         "TF", False),
    Cat2("TF02", "抽采管路",     "TF", False),
    Cat2("TF03", "密闭材料",     "TF", False),
    Cat2("TF04", "防灭火材料",   "TF", False),
    Cat2("TF99", "其他通防材料", "TF", False),
    # HX 化工材料 5 个 / 全部高敏感
    Cat2("HX01", "浮选药剂",     "HX", True),
    Cat2("HX02", "絮凝剂",       "HX", True),
    Cat2("HX03", "注浆材料",     "HX", True),
    Cat2("HX04", "阻燃材料",     "HX", True),
    Cat2("HX99", "其他化工材料", "HX", True),
    # GJ 工器具 7 个
    Cat2("GJ01", "手动工具",      "GJ", False),
    Cat2("GJ02", "电动工具",      "GJ", False),
    Cat2("GJ03", "气动工具",      "GJ", False),
    Cat2("GJ04", "测量量具",      "GJ", False),
    Cat2("GJ05", "便携检测仪表",  "GJ", False),
    Cat2("GJ06", "切削/钻磨工具", "GJ", False),
    Cat2("GJ99", "其他工器具",    "GJ", False),
    # LB 劳保用品 8 个
    Cat2("LB01", "安全帽",       "LB", False),
    Cat2("LB02", "矿灯",         "LB", False),
    Cat2("LB03", "自救器",       "LB", False),
    Cat2("LB04", "防护服装",     "LB", False),
    Cat2("LB05", "防护手套",     "LB", False),
    Cat2("LB06", "防护面具",     "LB", False),
    Cat2("LB07", "安全靴",       "LB", False),
    Cat2("LB99", "其他劳保用品", "LB", False),
    # BZ 包装材料 3 个
    Cat2("BZ01", "编织袋",    "BZ", False),
    Cat2("BZ02", "标签/标识", "BZ", False),
    Cat2("BZ99", "其他包装",  "BZ", False),
    # BG 办公用品 3 个
    Cat2("BG01", "纸张/打印耗材", "BG", False),
    Cat2("BG02", "文具用品",      "BG", False),
    Cat2("BG99", "其他办公用品",  "BG", False),
    # PS 排水材料 5 个（保留 / 暂停用）
    Cat2("PS01", "水泵配件",     "PS", False),
    Cat2("PS02", "排水管路",     "PS", False),
    Cat2("PS03", "阀门管件",     "PS", False),
    Cat2("PS04", "密封件",       "PS", False),
    Cat2("PS99", "其他排水材料", "PS", False),
]


# ============================================================================
# 数据：25 个常用计量单位
# ============================================================================
@dataclass(frozen=True)
class Unit:
    code: str
    name: str
    type_: str  # 长度 / 重量 / 体积 / 数量
    base: str
    factor: float


UNITS: List[Unit] = [
    # 长度 4
    Unit("M",  "米",   "长度", "M", 1),
    Unit("MM", "毫米", "长度", "M", 0.001),
    Unit("CM", "厘米", "长度", "M", 0.01),
    Unit("KM", "千米", "长度", "M", 1000),
    # 重量 3
    Unit("KG", "千克", "重量", "KG", 1),
    Unit("G",  "克",   "重量", "KG", 0.001),
    Unit("T",  "吨",   "重量", "KG", 1000),
    # 体积 3
    Unit("L",  "升",     "体积", "L", 1),
    Unit("ML", "毫升",   "体积", "L", 0.001),
    Unit("M3", "立方米", "体积", "L", 1000),
    # 数量 15
    Unit("件", "件", "数量", "件", 1),
    Unit("个", "个", "数量", "个", 1),
    Unit("套", "套", "数量", "套", 1),
    Unit("台", "台", "数量", "台", 1),
    Unit("张", "张", "数量", "张", 1),
    Unit("块", "块", "数量", "块", 1),
    Unit("根", "根", "数量", "根", 1),
    Unit("卷", "卷", "数量", "卷", 1),
    Unit("桶", "桶", "数量", "桶", 1),
    Unit("包", "包", "数量", "包", 1),
    Unit("盒", "盒", "数量", "盒", 1),
    Unit("瓶", "瓶", "数量", "瓶", 1),
    Unit("副", "副", "数量", "副", 1),
    Unit("双", "双", "数量", "双", 1),
    Unit("顶", "顶", "数量", "顶", 1),
]


# 防御性断言
assert len(CATEGORIES_L1) == 15, f"L1 应为 15，实为 {len(CATEGORIES_L1)}"
assert len(CATEGORIES_L2) == 95, f"L2 应为 95，实为 {len(CATEGORIES_L2)}"
assert len(UNITS) == 25, f"UNITS 应为 25，实为 {len(UNITS)}"


# ============================================================================
# 样式常量
# ============================================================================
THIN = Side(border_style="thin", color="B4B4B4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FONT = Font(name="微软雅黑", size=10)
ALIGN_LEFT = Alignment(vertical="center")
ALIGN_CENTER = Alignment(vertical="center", horizontal="center")
FILL_HI_SENS = PatternFill("solid", fgColor="FFF2CC")   # 浅黄：高敏感
FILL_INACTIVE = PatternFill("solid", fgColor="EAEAEA")  # 灰：保留/未启用

# 物料主数据 sheet 下拉验证覆盖范围（业务方可填的最大行数）
DV_RANGE_END = 2000

# 物料主数据 sheet 列定位（按 V0.2 表头顺序，从 1 开始）
COL_UNIT = "F"           # 第 6 列 = 计量单位
COL_CATEGORY = "G"       # 第 7 列 = V1.8 新分类编码（可留空）


def _dv_targets_our_columns(dv) -> bool:
    """判断 DV 是否作用于 F/G 列（用于幂等性清理）。"""
    try:
        ranges_str = str(dv.sqref)
    except Exception:
        return False
    return f"{COL_UNIT}2:" in ranges_str or f"{COL_CATEGORY}2:" in ranges_str


# ============================================================================
# 业务函数
# ============================================================================
def reorder_for_exclude_ps() -> tuple[List[Cat1], List[Cat2]]:
    """
    exclude-ps 模式下重排：PS 大类和 PS 子类整体移到末尾，
    下拉范围只截断到 PS 之前。
    """
    l1_active = [c for c in CATEGORIES_L1 if c.code != "PS"]
    l1_ps = [c for c in CATEGORIES_L1 if c.code == "PS"]

    l2_active = [c for c in CATEGORIES_L2 if c.parent != "PS"]
    l2_ps = [c for c in CATEGORIES_L2 if c.parent == "PS"]

    # 排序：14 启用大类 → 90 启用二级 → PS 大类 → 5 PS 二级
    return l1_active + l1_ps, l2_active + l2_ps


def fill_category_sheet(ws, exclude_ps: bool) -> tuple[int, int]:
    """
    写入「物料分类」sheet，返回 (last_data_row, dropdown_last_row)。
    - last_data_row：所有 110 行的最后一行（含 PS）
    - dropdown_last_row：下拉验证应覆盖的最后一行
        - exclude_ps=False：等于 last_data_row（PS 在下拉里）
        - exclude_ps=True： = last_data_row - 6（截断到 PS 之前）
    """
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    if exclude_ps:
        l1, l2 = reorder_for_exclude_ps()
    else:
        l1, l2 = CATEGORIES_L1, CATEGORIES_L2

    row = 2
    for c1 in l1:
        ws.cell(row=row, column=1, value=c1.code)
        ws.cell(row=row, column=2, value=c1.name)
        ws.cell(row=row, column=3, value=None)
        ws.cell(row=row, column=4, value=1)
        ws.cell(row=row, column=5, value=c1.high_sensitive)
        ws.cell(row=row, column=6, value=c1.is_active)
        row += 1

    for c2 in l2:
        is_active = c2.parent != "PS"
        ws.cell(row=row, column=1, value=c2.code)
        ws.cell(row=row, column=2, value=c2.name)
        ws.cell(row=row, column=3, value=c2.parent)
        ws.cell(row=row, column=4, value=2)
        ws.cell(row=row, column=5, value=c2.high_sensitive)
        ws.cell(row=row, column=6, value=is_active)
        row += 1

    last_data_row = row - 1
    dropdown_last_row = last_data_row - 6 if exclude_ps else last_data_row

    _apply_category_style(ws, last_data_row)
    return last_data_row, dropdown_last_row


def _apply_category_style(ws, last_row: int) -> None:
    for r in range(2, last_row + 1):
        is_hs = ws.cell(row=r, column=5).value is True
        is_inactive = ws.cell(row=r, column=6).value is False
        for c in range(1, 7):
            cell = ws.cell(row=r, column=c)
            cell.font = FONT
            cell.border = BORDER
            cell.alignment = ALIGN_CENTER if c in (1, 3, 4, 5, 6) else ALIGN_LEFT
            if is_inactive:
                cell.fill = FILL_INACTIVE
            elif is_hs:
                cell.fill = FILL_HI_SENS
    for col, width in zip("ABCDEF", [14, 32, 16, 12, 14, 12]):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"


def fill_unit_sheet(ws) -> int:
    """写入「计量单位」sheet，返回最后一行行号。"""
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    row = 2
    for u in UNITS:
        ws.cell(row=row, column=1, value=u.code)
        ws.cell(row=row, column=2, value=u.name)
        ws.cell(row=row, column=3, value=u.type_)
        ws.cell(row=row, column=4, value=u.base)
        ws.cell(row=row, column=5, value=u.factor)
        row += 1
    last_row = row - 1

    for r in range(2, last_row + 1):
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            cell.font = FONT
            cell.border = BORDER
            cell.alignment = ALIGN_CENTER
    for col, width in zip("ABCDE", [12, 16, 12, 16, 14]):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    return last_row


def add_material_dropdowns(ws_mat, cat_dropdown_last_row: int, unit_last_row: int) -> None:
    """给「物料主数据」sheet 的 unit/category_code 列加 Excel 下拉验证。

    幂等保证：先清空已有的 list 类型 DataValidation（其他类型如 textLength 等不动），
    避免重复运行脚本时 DV 叠加产生重复规则。
    """
    existing = list(ws_mat.data_validations.dataValidation)
    ws_mat.data_validations.dataValidation = [
        dv for dv in existing
        if dv.type != "list" or not _dv_targets_our_columns(dv)
    ]

    # Excel/WPS 要求 dataValidation 的 formula1 **不带前导 `=`**；
    # 如果带了 `=`，xml 里会落成 `<formula1>=计量单位!...</formula1>`，
    # Excel/WPS 解析失败 → 打开时静默丢弃 DV 甚至触发"修复"回退数据。
    # 参考 openpyxl 文档：DataValidation.formula1 期望的是裸公式字符串。
    dv_unit = DataValidation(
        type="list",
        formula1=f"计量单位!$A$2:$A${unit_last_row}",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="计量单位无效",
        error="请从「计量单位」sheet 的标准单位中选择",
        showInputMessage=True,
        promptTitle="计量单位",
        prompt="从下拉中选择标准单位（参考「计量单位」sheet）",
    )
    dv_unit.add(f"{COL_UNIT}2:{COL_UNIT}{DV_RANGE_END}")
    ws_mat.add_data_validation(dv_unit)

    dv_cat = DataValidation(
        type="list",
        formula1=f"物料分类!$A$2:$A${cat_dropdown_last_row}",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="分类编码无效",
        error="请从「物料分类」sheet 范围内选择；不在 V1.8 范围内需走《物资编码规范文档》§六分类提报机制",
        showInputMessage=True,
        promptTitle="V1.8 新分类编码",
        prompt="二级分类编码（如 ZH01 / JD01 / HG01），从下拉选择。可留空，系统按 M-18 自动映射",
    )
    dv_cat.add(f"{COL_CATEGORY}2:{COL_CATEGORY}{DV_RANGE_END}")
    ws_mat.add_data_validation(dv_cat)


# ============================================================================
# CLI
# ============================================================================
def find_repo_root(start: Path) -> Path:
    """从脚本位置向上找仓库根（含 .git 目录）。"""
    p = start.resolve()
    for parent in [p, *p.parents]:
        if (parent / ".git").exists():
            return parent
    return start.parent


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="02 物资主数据 xlsx 灌库脚本（V1.8 权威基线 + 25 标准单位 + 下拉验证）",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "--file",
        type=str,
        default=None,
        help=f"目标 xlsx 路径（默认：仓库内 {DEFAULT_REL_PATH}）",
    )
    parser.add_argument(
        "--exclude-ps",
        action="store_true",
        help="PS 保留大类完全不进下拉（PS 区移到 sheet 末尾，下拉范围截断到 PS 之前）",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="不写盘，只打印将要做的事",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()

    if args.file:
        target = Path(args.file).expanduser().resolve()
    else:
        repo_root = find_repo_root(Path(__file__))
        target = repo_root / DEFAULT_REL_PATH

    print(f"[fill_material_template] 目标文件：{target}")
    print(f"[fill_material_template] PS 保留大类：{'排除出下拉' if args.exclude_ps else '保留在下拉（灰底标 is_active=false）'}")
    print(f"[fill_material_template] dry-run：{args.dry_run}")

    if not target.exists():
        print(f"错误：文件不存在：{target}", file=sys.stderr)
        return 2

    # Excel/WPS 锁文件检测
    lock = target.parent / f".~{target.name}"
    if lock.exists():
        print(f"错误：检测到锁文件 {lock.name}，请先关闭 Excel/WPS 客户端再重跑", file=sys.stderr)
        return 3

    # ---- 计算预期行数（用于 dry-run 输出 + 后续校验）----
    expected_cat_rows = len(CATEGORIES_L1) + len(CATEGORIES_L2)  # 110
    expected_unit_rows = len(UNITS)                              # 25
    expected_dropdown_last = expected_cat_rows + 1 - 6 if args.exclude_ps else expected_cat_rows + 1
    # 行号从 2 起算：last_data_row = 1 + 110 = 111；exclude_ps 时 dropdown_last = 105
    cat_data_last_row = 1 + expected_cat_rows
    cat_dropdown_last_row = cat_data_last_row - 6 if args.exclude_ps else cat_data_last_row

    print()
    print(f"  ▸ 物料分类 sheet：写入 {expected_cat_rows} 行（{len(CATEGORIES_L1)} 大类 + {len(CATEGORIES_L2)} 二级）")
    print(f"  ▸ 计量单位 sheet：写入 {expected_unit_rows} 行")
    print(f"  ▸ 物料主数据 sheet：")
    print(f"      F 列 unit 下拉范围      = 计量单位!$A$2:$A${1 + expected_unit_rows}（formula1 不带前导 =）")
    print(f"      G 列 category 下拉范围  = 物料分类!$A$2:$A${cat_dropdown_last_row}（formula1 不带前导 =）")
    print(f"      下拉覆盖行范围          = R2:R{DV_RANGE_END}")
    print()

    if args.dry_run:
        print("[dry-run] 已计算所有写入计划但未落盘。")
        return 0

    # ---- 实际写入 ----
    try:
        wb = load_workbook(target)
    except Exception as e:
        print(f"错误：无法打开 xlsx：{e}", file=sys.stderr)
        print("       常见原因：文件被 Excel/WPS 锁定 / 不是有效 xlsx", file=sys.stderr)
        return 4

    for required_sheet in ("物料分类", "计量单位", "物料主数据"):
        if required_sheet not in wb.sheetnames:
            print(f"错误：xlsx 缺少必需 sheet「{required_sheet}」（实际 sheets: {wb.sheetnames}）", file=sys.stderr)
            return 5

    cat_last, cat_dropdown_last = fill_category_sheet(wb["物料分类"], exclude_ps=args.exclude_ps)
    unit_last = fill_unit_sheet(wb["计量单位"])
    add_material_dropdowns(wb["物料主数据"], cat_dropdown_last, unit_last)

    wb.save(target)
    wb.close()

    print(f"✓ 写入完成")
    print(f"  - 物料分类 sheet：R2-R{cat_last}（含 PS 行；下拉截至 R{cat_dropdown_last}）")
    print(f"  - 计量单位 sheet：R2-R{unit_last}")
    print(f"  - 物料主数据 sheet：F/G 列下拉验证已生效（R2:R{DV_RANGE_END}）")
    return 0


if __name__ == "__main__":
    sys.exit(main())
